from datetime import datetime
from django.http import JsonResponse, HttpResponse
from rest_framework import generics
from .models import *
import base64
from django.http import HttpResponse
from django.conf import settings
import os
from django.contrib.auth.hashers import make_password, check_password
from openpyxl import load_workbook

#El test de la api funcional
class Test(generics.GenericAPIView):
    def get(self, request):
        response = {
            'response': "Hello world! Test api"
        }
        return JsonResponse(response, status=200)
    
class ValidateLogin(generics.GenericAPIView):
    def get(self, request, cedula, passw):
        try:
            worker=Usuario.objects.get(cedula=cedula)
            workerPass=worker.password
            if check_password(passw, workerPass):
                return JsonResponse({"Cedula":worker.cedula},status=200)
            else:
                return HttpResponse({'Credenciales incorrectas'},status=500)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)
        
class TipoUsuario(generics.GenericAPIView):
    def get(self, request, cedula):
        try:
            worker=Usuario.objects.get(cedula=cedula)
            return JsonResponse({"Rol":worker.rol},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)

class CargarInfoUsuario(generics.GenericAPIView):
    def get(self, request, cedula):
        try:
            genero=''
            worker=Usuario.objects.get(cedula=cedula)
            if(worker.genero=="M"):
                genero="Masculino"
            elif(worker.genero=="F"):
                genero="Femenino"
            return JsonResponse({"Apellidos":worker.apellidos,"Nombres":worker.nombres,"Cedula":worker.cedula,"Correo":worker.email,"Nacimiento":worker.fechaNacimiento,"Contrato":worker.fechaContrato,"Genero":genero},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)
        
class CargarDatosTablaAdmin(generics.GenericAPIView):
    def get(self, request, cedula, mes, year):
        try:
            list_marcas=list()
            datos=list()
            isPublished=True
            worker=Usuario.objects.get(cedula=cedula)
            biometrico=worker.codigoBiometrico
            marcaciones=Marcacion.objects.filter(codigoBiometrico=biometrico)
            for marcacion in marcaciones:
                if(marcacion.publicado==False):
                    isPublished=False
                mesMarca=marcacion.fecha.strftime('%m')
                yearMarca=marcacion.fecha.strftime('%y')
                if(mes==mesMarca and year=="20"+yearMarca):
                    datos.append(marcacion.fecha)
                    datos.append(marcacion.horaEntrada)
                    datos.append(marcacion.horaSalida)
                    ent=marcacion.horaEntrada
                    h=ent.hour
                    m=str(ent.minute)
                    s=str(ent.second)
                    diffH= h-8
                    if(diffH<0):
                        atraso='00:00:00'
                    else:
                        atraso=str(diffH)+":"+m+":"+s
                    datos.append(atraso)
                    list_marcas.append(datos)
                    datos=list()
            return JsonResponse({"Marcas":list_marcas,"Pendientes":isPublished},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)

class CargarSolicitudes(generics.GenericAPIView):
    def get(self, request,cedula):
        solicitud=list()
        worker=Usuario.objects.get(cedula=cedula)
        tiposJust=TipoJustificacion.objects.filter(codigoBiometrico=worker.codigoBiometrico)
        for justi in tiposJust:
            descripcion=justi.descripcion
            idTipo=justi.idTipo
            incons=Inconsistencia.objects.get(idTipo=idTipo)
            id=incons.idMarcacion.idMarcacion
            marcacion=Marcacion.objects.get(idMarcacion=id)
            estado=''
            if(incons.estado=="P"):
                user=Usuario.objects.get(codigoBiometrico=marcacion.codigoBiometrico.codigoBiometrico)
                if(incons.estado=="P"):
                    estado='Pendiente'
                solicitud.append({'id':idTipo,'fecha':marcacion.fecha,'descripcion':descripcion,'estado':estado})
        return JsonResponse({"justificaciones":solicitud},status=200)
        
class EnviarSolicitud(generics.GenericAPIView):
    #Necesito: biometrico de quien modifica,descripcion,fechaIngreso para tipo just
    #Necesito: idTipo,idMarcacion para la inconsistencia
    def post(self,request):
        cedula=request.data["cedula"]
        fecha=request.data["fecha"]
        descripcion=request.data["motivo"]
        fechaA=datetime.today().strftime('%Y-%m-%d')
        try:
            user=Usuario.objects.get(cedula=cedula)
            try:
                marcacion=Marcacion.objects.get(fecha=fecha,codigoBiometrico=user)
                incons=Inconsistencia.objects.filter(idMarcacion=marcacion.idMarcacion)
                if(incons.count()==0):
                    justif=TipoJustificacion(codigoBiometrico=user,descripcion=descripcion,fechaIngreso=fechaA)
                    justif.save()
                    incons=Inconsistencia(idTipo=justif,idMarcacion=marcacion,estado="P")
                    incons.save()
                    return JsonResponse({"msg":"Solicitud enviada"},status=200)
                else:
                    return JsonResponse({"msg":"Ya existe una justificacion de este tipo"},status=500)
            except(Marcacion.objects.get(fecha=fecha,codigoBiometrico=user).DoesNotExist):
                return JsonResponse({'msg:':{"No se encontró la marcacion"}},status=500)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg:':{"No se encontró al usuario"}},status=500)
        
class NewPassword(generics.GenericAPIView):
    def post(self,request,cedula):
        try:
            passOld=request.data['old']
            passNew=request.data['new']
            user=Usuario.objects.get(cedula=cedula)
            if(check_password(passOld,user.password)):
                user.password=make_password(passNew)
                user.save()
                return JsonResponse({"msg":"Contraseña cambiada correctamente"},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg:':{"No se encontró al usuario"}},status=404)

class GuardarCambioHoras(generics.GenericAPIView):
    def post(self,request,cedula):
        fecha=request.data['fecha']
        horaE=request.data['horaE']
        horaS=request.data['horaS']
        fechaN=datetime.strptime(fecha,'%Y-%m-%d').date()
        print("fecha",fechaN)
        print("entr",horaE)
        print("salida",horaS)
        user=Usuario.objects.get(cedula=cedula)
        try:
            marcacion=Marcacion.objects.get(codigoBiometrico=user.codigoBiometrico,fecha=fechaN)
            if(horaE!='0'):
                marcacion.horaEntrada=horaE
            if(horaS!='0'):
                marcacion.horaSalida=horaS
            marcacion.save()
            return JsonResponse({"msg":"Los horarios fueron cambiados con éxito"},status=200)
        except(Marcacion.objects.get(codigoBiometrico=user.codigoBiometrico,fecha=fechaN).DoesNotExist):
            return JsonResponse({"msg":"No hay marcacion disponible"},status=404)

class CargarSolicitudesAdmin(generics.GenericAPIView):
    def get(self,request):
        solicitud=list()
        tiposJust=TipoJustificacion.objects.all()
        for justi in tiposJust:
            descripcion=justi.descripcion
            idTipo=justi.idTipo
            incons=Inconsistencia.objects.get(idTipo=idTipo)
            id=incons.idMarcacion.idMarcacion
            marcacion=Marcacion.objects.get(idMarcacion=id)
            estado=''
            if(incons.estado=="P"):
                user=Usuario.objects.get(codigoBiometrico=marcacion.codigoBiometrico.codigoBiometrico)
                if(incons.estado=="P"):
                    estado='Pendiente'
                solicitud.append({'id':idTipo,'worker':user.nombres+" "+user.apellidos,'fecha':marcacion.fecha,'descripcion':descripcion,'estado':estado})
        return JsonResponse({"justificaciones":solicitud},status=200)

class CargarWorkers(generics.GenericAPIView):
    def get(self, request):
        workers=list()
        usuarios=Usuario.objects.all()
        for user in usuarios:
            workers.append({"value":user.cedula,"label":user.apellidos+" "+user.nombres})
        return JsonResponse({"workers":workers},status=200)
    
class HistoricoSolicitudesAdmin(generics.GenericAPIView):
    def get(self,request):
        solicitud=list()
        tiposJust=TipoJustificacion.objects.all()
        idCont=0
        for justi in tiposJust:
            descripcion=justi.descripcion
            idTipo=justi.idTipo
            incons=Inconsistencia.objects.get(idTipo=idTipo)
            id=incons.idMarcacion.idMarcacion
            marcacion=Marcacion.objects.get(idMarcacion=id)
            estado=''
            if(incons.estado=="NJ" or incons.estado=="J"):
                user=Usuario.objects.get(codigoBiometrico=marcacion.codigoBiometrico.codigoBiometrico)
                if(incons.estado=="NJ"):
                    estado='Negada'
                elif(incons.estado=='J'):
                    estado='Justificada'
                solicitud.append({'id':idCont,'worker':user.nombres+" "+user.apellidos,'fecha':marcacion.fecha,'descripcion':descripcion,'estado':estado})
            idCont=idCont+1
        return JsonResponse({"justificaciones":solicitud},status=200)

class HistoricoSolicitudes(generics.GenericAPIView):
    def get(self, request,cedula):
        solicitud=list()
        idCont=0
        worker=Usuario.objects.get(cedula=cedula)
        tiposJust=TipoJustificacion.objects.filter(codigoBiometrico=worker.codigoBiometrico)
        for justi in tiposJust:
            descripcion=justi.descripcion
            idTipo=justi.idTipo
            incons=Inconsistencia.objects.get(idTipo=idTipo)
            id=incons.idMarcacion.idMarcacion
            marcacion=Marcacion.objects.get(idMarcacion=id)
            estado=''
            if(incons.estado=="NJ" or incons.estado=="J"):
                user=Usuario.objects.get(codigoBiometrico=marcacion.codigoBiometrico.codigoBiometrico)
                if(incons.estado=="NJ"):
                    estado='Negada'
                elif(incons.estado=='J'):
                    estado='Justificada'
                solicitud.append({'id':idCont,'fecha':marcacion.fecha,'descripcion':descripcion,'estado':estado})
            idCont=idCont+1
        return JsonResponse({"justificaciones":solicitud},status=200)

class CargarDatosTabla(generics.GenericAPIView):
    def get(self, request, cedula, mes, year):
        try:
            list_marcas=list()
            datos=list()
            worker=Usuario.objects.get(cedula=cedula)
            biometrico=worker.codigoBiometrico
            marcaciones=Marcacion.objects.filter(codigoBiometrico=biometrico,publicado=True)
            for marcacion in marcaciones:
                mesMarca=marcacion.fecha.strftime('%m')
                yearMarca=marcacion.fecha.strftime('%y')
                if(mes==mesMarca and year=="20"+yearMarca):
                    datos.append(marcacion.fecha)
                    datos.append(marcacion.horaEntrada)
                    datos.append(marcacion.horaSalida)
                    ent=marcacion.horaEntrada
                    h=ent.hour
                    m=str(ent.minute)
                    s=str(ent.second)
                    diffH= h-8
                    if(diffH<0):
                        atraso='00:00:00'
                    else:
                        atraso=str(diffH)+":"+m+":"+s
                    datos.append(atraso)
                    list_marcas.append(datos)
                    datos=list()
            return JsonResponse({"Marcas":list_marcas},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)
        
class Publicar(generics.GenericAPIView):
    def post(self, request):
        try:
            cedula=request.data["cedula"]
            mes=request.data["mes"]
            year=request.data["year"]
            worker=Usuario.objects.get(cedula=cedula)
            biometrico=worker.codigoBiometrico
            marcaciones=Marcacion.objects.filter(codigoBiometrico=biometrico)
            for marcacion in marcaciones:
                mesMarca=marcacion.fecha.strftime('%m')
                yearMarca=marcacion.fecha.strftime('%y')
                if(mes==mesMarca and year=="20"+yearMarca):
                    marcacion.publicado=True
                    marcacion.save()
            return JsonResponse({"Publicado":mes+" "+year},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)
        
class ManejarSolicitud(generics.GenericAPIView):
    def post(self,request):
        try:
            idTipo=request.data["idTipo"]
            orden=request.data["orden"]
            incons=Inconsistencia.objects.get(idTipo=idTipo)
            if(orden=="aceptar"):
                incons.estado="J"
            else:
                incons.estado="NJ"
            incons.save()
            return JsonResponse({"Solicitud":orden},status=200)
        except(Inconsistencia.objects.get(idTipo=idTipo).DoesNotExist):
            return JsonResponse({'msg': 'No existe una inconsistencia con ese id'},status=404)
        
class Registrar(generics.GenericAPIView):
    def post(self,request):
        cedula=request.data["cedula"]
        biom=request.data["biom"]
        nombres=request.data["nombres"]
        apellidos=request.data["apellidos"]
        email=request.data["email"]
        fechaB = datetime.strptime(request.data["born"], "%a, %d %b %Y %H:%M:%S %Z")
        born = fechaB.strftime("%Y-%m-%d")
        fechaC = datetime.strptime(request.data["cont"], "%a, %d %b %Y %H:%M:%S %Z")
        cont = fechaC.strftime("%Y-%m-%d")
        gen=request.data["gen"]
        users=Usuario.objects.filter(cedula=cedula)
        hash = make_password(cedula)
        print(cedula,biom,born,email,cont,nombres,apellidos,gen)
        if(users.count()==0):
            if(Usuario.objects.filter(codigoBiometrico=biom).count()==0):
                newUser=Usuario(cedula=cedula,codigoBiometrico=biom,rol='USER',fechaNacimiento=born,email=email,password=hash,fechaContrato=cont,nombres=nombres,apellidos=apellidos,genero=gen)
                newUser.save()
                return JsonResponse({'msg': 'Usuario registrado'},status=200)
        else:
            return JsonResponse({'msg': 'Usuario ya existia'},status=500)
        
class MarcacionIndividual(generics.GenericAPIView):
    #necesito biometrico, fecha y horas de entrada y salida
    def post(self,request):
        cedula=request.data["cedula"]
        fecha=request.data["fecha"]
        horaE=request.data["horaE"]
        horaS=request.data["horaS"]
        try:
            user=Usuario.objects.get(cedula=cedula)
            marca=Marcacion(codigoBiometrico=user,fecha=fecha,horaEntrada=horaE,horaSalida=horaS)
            marca.save()
            return JsonResponse({'msg': 'marcacion registrado'},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario no encontrado'},status=500)
        
class SubirImg(generics.GenericAPIView):
    def post(self,request):
        img=request.FILES['image']
        cedula=request.data["cedula"]
        try:
            user=Usuario.objects.get(cedula=cedula)
            with open('media/' + img.name, 'wb+') as destination:
                for chunk in img.chunks():
                    destination.write(chunk)
            user.nombreimg=img.name
            user.save()
            return JsonResponse({'message': 'Imagen subida exitosamente'})
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'message': 'No se pudo subir la imagen'})
    
class ObtenerImg(generics.GenericAPIView):
    def get(self,request,cedula):
        try:
            user=Usuario.objects.get(cedula=cedula)
            rutaimg=user.nombreimg
            if(rutaimg==None):
                rutaimg="picture-placeholder.jpg"
            route=os.path.join(settings.MEDIA_ROOT, rutaimg)
            try:
                with open(f'media/{rutaimg}', 'rb') as imagen_file:
                    imagen_base64 = base64.b64encode(imagen_file.read()).decode('utf-8')
                    return JsonResponse({'imagen': imagen_base64})
            except FileNotFoundError:
                return HttpResponse("La imagen no fue encontrada", status=404)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'message': 'No se pudo encontrar la ruta de la imagen'})
        
        
class IsPublicado(generics.GenericAPIView):
    def get(self, request,cedula,mes,year):
        publicado=True
        try:
            worker=Usuario.objects.get(cedula=cedula)
            biometrico=worker.codigoBiometrico
            marcaciones=Marcacion.objects.filter(codigoBiometrico=biometrico)
            for marcacion in marcaciones:
                mesMarca=marcacion.fecha.strftime('%m')
                yearMarca=marcacion.fecha.strftime('%y')
                if(mes==mesMarca and year=="20"+yearMarca):
                    if(marcacion.publicado==False):
                        publicado=False
            return JsonResponse({"isPublicado":publicado},status=200)
        except(Usuario.objects.get(cedula=cedula).DoesNotExist):
            return JsonResponse({'msg': 'Usuario vinculado a esta cédula no existe'},status=404)
        
class MarcarArchivo(generics.GenericAPIView):
    def post(self,request):
        if request.method == 'POST':
            archivo = request.data['archivo']
            ruta_destino = os.path.join(settings.MEDIA_ROOT, archivo.name)
            with open(ruta_destino, 'wb') as destino:
                for chunk in archivo.chunks():
                    destino.write(chunk)

            wb = load_workbook(archivo, read_only=True)
            hoja = wb.active
            for fila in hoja.iter_rows(min_row=3, values_only=True):  
                # Aqui corresponde el manejo del guardo de informacion del archivo
                for celda in fila:
                    print(celda)
            wb.close()

            return JsonResponse({'mensaje': 'datos guardados correctamente.'})
        return JsonResponse({'mensaje': 'Error al procesar la solicitud.'}, status=400)


class Hashing(generics.GenericAPIView):
    def get(self,request,passw):
        passwd=make_password(passw)
        return JsonResponse({'hashed': passwd})